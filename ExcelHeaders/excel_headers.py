
from openpyxl import load_workbook

print('starting ...\n')


wb = load_workbook(filename = 'ExcelTest1.xlsx')
sheet_names = wb.sheetnames
print(sheet_names)
sheet1_name = sheet_names[0]
print('first sheet name=', sheet1_name)

for sheet_name in sheet_names:
    print('\nCurrent sheet name: ', sheet_name)

for ws in wb:
    print('\nCurrent title: ', ws.title)
    is_header_row_found = False
    row_num = 0
    print("  headers for this sheet ...")
    for row in ws.iter_rows(min_row=1, max_col=100, max_row=20, values_only=True):
        row_num = row_num + 1
        # print("  row number: ", row_num)
        if is_header_row_found:
            break;
        for cell_value in row:
            # print("  cell value:", cell_value)
            if cell_value != None:
                is_header_row_found = True
                print("    ", cell_value)


print('\nDone.\n')

